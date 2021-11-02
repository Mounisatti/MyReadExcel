import openpyxl

wk = openpyxl.load_workbook("C://Users//user//Desktop//ReadExcel.xlsx")
def  Rows_of_Numbers(Sheetname):
     sh = wk[Sheetname]
     #print(sh.max_row)
     return sh.max_row

def  Cell_data(Sheetname, row, cell):
     sh = wk[Sheetname]
     cell = sh.cell(int(row), int(cell))
     return cell.value


print(Rows_of_Numbers("Sheet1"))
print(Cell_data("Sheet1",1,1))
#sh = wk["Sheet1"]


#print(cell.value)